"""
Offline test for the Plant-to-Depot block parser — no network/DB dependency.
Loads the real dummy workbook directly via openpyxl (simulating the grid shape
the Sheets API's values.batchGet would return) and runs the same parser that
the live sync endpoint will use.
"""
import sys; sys.path.insert(0, '.')
import openpyxl
from services.sales_sync import parse_tab_title, parse_month_tab


def sheet_category_total(grid):
    """
    Cross-check helper (test-only, not part of the real parser): every category
    block's own 'Grand Total'/'TOTAL' row has its amount/value 3 columns right of
    the label, in every block (Seat Cover, Accessories, Mats, Electronics) —
    confirmed by inspection. These are Excel SUM-range formulas, so (unlike the
    sheet's bottom unlabeled per-depot block, which turned out in FEB-26 to be
    built from fixed cell references and silently missed a mis-placed value) they
    stay correct even when a human typed a number on the wrong row. Summing all
    of them except the bottom-most one (which IS that unreliable unlabeled block,
    and is deliberately excluded the same way the real parser excludes it) gives
    an independent total to validate our own row-by-row summed amounts against.
    """
    cells = []
    for r_idx, line in enumerate(grid, start=1):
        for c_idx, v in enumerate(line, start=1):
            if isinstance(v, str) and v.strip().upper() in ("GRAND TOTAL", "TOTAL"):
                cells.append((r_idx, c_idx))
    if not cells:
        return None
    bottom_row = max(r for r, _ in cells)
    total = 0.0
    for r, c in cells:
        if r == bottom_row:
            continue  # the unlabeled per-depot block — excluded, same as the real parser
        v = line_value(grid, r, c + 3)
        if isinstance(v, (int, float)):
            total += v
    return round(total, 2)


def line_value(grid, row, col):
    r, c = row - 1, col - 1
    if r < 0 or r >= len(grid):
        return None
    line = grid[r]
    if c < 0 or c >= len(line):
        return None
    return line[c]

WORKBOOK = r"D:\MIS\Local_sheets\Sales\Plant to Depot-MIS DATA 26.xlsx"


def sheet_to_grid(ws):
    grid = []
    for row in ws.iter_rows():
        grid.append([c.value for c in row])
    return grid


def main():
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)

    all_records, all_errors = [], []
    month_totals = {}

    for title in wb.sheetnames:
        ym = parse_tab_title(title)
        assert ym is not None, f"tab title '{title}' should be recognized as a month"
        year, month = ym
        grid = sheet_to_grid(wb[title])
        records, errors = parse_month_tab(year, month, grid, tab_title=title)
        all_records += records
        all_errors += errors
        month_totals[title] = round(sum(r["amount"] for r in records), 2)
        cross_check = sheet_category_total(grid)
        match = "OK" if cross_check is not None and abs(cross_check - month_totals[title]) < 0.01 else "MISMATCH"
        print(f"{title:10s} -> {len(records):2d} records, {len(errors)} messages, "
              f"our total = {month_totals[title]:,.2f}, sum of sheet's category Grand Totals = {cross_check}, [{match}]")
        assert match == "OK", f"{title}: parsed total {month_totals[title]} != sheet's category-total sum {cross_check}"

    print(f"\nTotal records across all tabs: {len(all_records)}")
    if all_errors:
        print("Messages:")
        for e in all_errors:
            print(" ", e)

    # ── Hand-verified assertions (from manual cell inspection) ──────────────
    def find(records, year, month, depot, brand, category):
        for r in records:
            if (r["sale_year"], r["sale_month"], r["depot"], r["brand"], r["category"]) == (year, month, depot, brand, category):
                return r
        return None

    jan_af_seatcover_janak = find(all_records, 2026, 1, "Janak Motors", "Autoform", "Seat Cover")
    assert jan_af_seatcover_janak is not None, "JAN-26 Janak/Autoform/Seat Cover record missing"
    assert jan_af_seatcover_janak["amount"] == 12828725.13, jan_af_seatcover_janak

    jan_united_ac_seatcover = find(all_records, 2026, 1, "United Auto", "Autocruze", "Seat Cover")
    assert jan_united_ac_seatcover is not None, "JAN-26 United Auto/Autocruze/Seat Cover record missing"
    expected_united_ac_amount = round(167489.0 + 1566980.08, 2)  # Chandigarh + Ludhiana summed
    assert jan_united_ac_seatcover["amount"] == expected_united_ac_amount, jan_united_ac_seatcover

    assert month_totals["JAN-26"] == 66607700.22, month_totals["JAN-26"]
    assert month_totals["MAR 26"] == 57591810.55, month_totals["MAR 26"]

    print("\nAll assertions passed.")


if __name__ == "__main__":
    main()
