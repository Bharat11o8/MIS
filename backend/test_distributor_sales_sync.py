"""
Offline test for the Depot-to-Distributor parser — no network/DB dependency.
Loads the real quarterly workbook directly via openpyxl (simulating the grid
shape the Sheets API's values.get would return) and runs the same parser that
the live sync endpoint will use.
"""
import sys; sys.path.insert(0, '.')
import openpyxl
from services.distributor_sales_sync import parse_distributor_grid

WORKBOOK = r"D:\MIS\Local_sheets\Sales\MIS -AFAC-After Market -Distributor_Sales Team -Apr to Jun Movement (FY1)-2026.xlsx"


def sheet_to_grid(ws):
    return [[c.value for c in row] for row in ws.iter_rows()]


def main():
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    ws = wb[wb.sheetnames[0]]
    grid = sheet_to_grid(ws)

    records, errors = parse_distributor_grid(grid, calendar_year=2026)
    print(f"{len(records)} records, {len(errors)} messages")
    for e in errors:
        print(" ", e)

    def total_for(name):
        return round(sum(r["amount"] for r in records if r["distributor"] == name and r["entity_type"] == "distributor"), 2)

    def target_for(name):
        rec = next(r for r in records if r["distributor"] == name and r["entity_type"] in ("distributor", "depot_direct"))
        return rec["target"]

    def entity_type_for(name):
        return {r["entity_type"] for r in records if r["distributor"] == name}

    def area_head_total(area_head, month, category):
        return next(
            r["amount"] for r in records
            if r["entity_type"] == "area_head_total" and r["area_head"] == area_head
            and r["sale_month"] == month and r["category"] == category
        )

    def grand_total(month, category):
        return next(
            r["amount"] for r in records
            if r["entity_type"] == "grand_total" and r["sale_month"] == month and r["category"] == category
        )

    # ── ADITYA ENTERPRISES (hand-verified) ──────────────────────────────────
    aditya_target = target_for("ADITYA ENTERPRISES")
    aditya_total = total_for("ADITYA ENTERPRISES")
    assert aditya_target == 14662500, aditya_target
    assert aditya_total == 11460872.0, aditya_total
    aditya_pct = aditya_total / aditya_target * 100
    assert abs(aditya_pct - 78.16451492) < 0.01, aditya_pct

    # ── ANIL SHARMA group's own TOTAL row — no adjustment, matches sum of its
    # 2 distributors exactly (verified independently above: 4221125+4554817=8775942)
    anil_total_target = next(
        r["target"] for r in records if r["entity_type"] == "area_head_total" and r["area_head"] == "ANIL SHARMA"
    )
    assert anil_total_target == 34912500, anil_total_target
    assert area_head_total("ANIL SHARMA", 4, "SAM") == 8775942, area_head_total("ANIL SHARMA", 4, "SAM")

    # ── JANAK MOTORS — depot_direct, no target, negative May SAM ────────────
    assert entity_type_for("JANAK MOTORS") == {"depot_direct"}, entity_type_for("JANAK MOTORS")
    assert target_for("JANAK MOTORS") is None
    janak_may_sam = next(
        r["amount"] for r in records
        if r["distributor"] == "JANAK MOTORS" and r["sale_month"] == 5 and r["category"] == "SAM"
    )
    assert janak_may_sam == -2992906, janak_may_sam

    # ── ATUL DWIVEDI's own TOTAL row carries a real ~7.9L manual adjustment on
    # Apr SAM with no corresponding distributor row (3,892,049 vs. its 4
    # distributor rows summing to only 3,099,176) — confirmed with the user this
    # is an intentional business adjustment, not a sheet error. We take the
    # sheet's own TOTAL row as authoritative rather than recomputing it away.
    atul_distributor_sum_apr_sam = round(sum(
        r["amount"] for r in records
        if r["entity_type"] == "distributor" and r["area_head"] == "ATUL DWIVEDI"
        and r["sale_month"] == 4 and r["category"] == "SAM"
    ), 2)
    assert atul_distributor_sum_apr_sam == 3099176.0, atul_distributor_sum_apr_sam
    assert area_head_total("ATUL DWIVEDI", 4, "SAM") == 3892049, area_head_total("ATUL DWIVEDI", 4, "SAM")

    # ── GRAND TOTAL row taken as-is — now matches the sheet's own printed cell
    # exactly (48,784,278), since it's read directly rather than recomputed.
    assert grand_total(4, "SAM") == 48784278, grand_total(4, "SAM")
    grand_total_target = next(r["target"] for r in records if r["entity_type"] == "grand_total")
    assert grand_total_target == 225000000, grand_total_target

    print("\nAll assertions passed.")


if __name__ == "__main__":
    main()
